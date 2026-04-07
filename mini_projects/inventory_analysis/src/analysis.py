import argparse
import pandas as pd
from pathlib import Path
from openpyxl.styles import Font,PatternFill,Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule


pd.options.display.float_format='{:,.0f}'.format


def load_data(filepath: str) -> pd.DataFrame:
    """load raw excel data."""
    return pd.read_excel(filepath)

def clean_data(df: pd.DataFrame) -> pd.DataFrame:
    """Clean raw data and standardize columns."""
    df=df.copy()

    # standardize column names
    df.columns = df.columns.str.strip()
    # drop rows missing key business fields
    df = df.dropna(subset=["Dept","SKU","Item"])
    # Drop unsuded column if present
    df = df.drop(columns= ["Metrics"], errors="ignore")
    # ensure data columns are datetime
    for col in ["Week Start", "Week End"]:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")

    # reset index after cleaning
    df = df.reset_index(drop=True)

    return df

def validate_columns(df:pd.DataFrame) -> None:
    """Validate required columns exist before analysis."""
    required_cols = [
        "Dept",
        "Class",
        "Subclass",
        "SKU",
        "Item",
        "Week Start",
        "Week End",
        "DTC Netsales $s",
        "DTC Netsales $s LY"
    ]

    missing = [col for col in required_cols if col not in df.columns]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")


def get_demand_column(df: pd.DataFrame) -> str:
    """Return the correct demand column name from possible variants."""
    possible_cols = ["DTC Gross Demand $s", "DTC Gross Demand Us"]
    for col in possible_cols:
        if col in df.columns:
            return col
    raise ValueError(f"Could not find a demand column. Checked: {possible_cols}")


def get_latest_week(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.Timestamp]:
    """Filter to the latest week in the dataset."""
    latest_week = df["Week End"].max()
    df_latest = df[df["Week End"] == latest_week].copy()
    return df_latest, latest_week


def get_prior_week(df: pd.DataFrame, latest_week: pd.Timestamp) -> tuple[pd.DataFrame, pd.Timestamp]:
    """Filter to the prior week before latest week."""
    prior_week = df[df["Week End"] < latest_week]["Week End"].max()
    df_prior = df[df["Week End"] == prior_week].copy()
    return df_prior, prior_week


def build_category_summary(df: pd.DataFrame, demand_col: str) -> pd.DataFrame:
    """Build category-level sales summary with YoY %."""
    summary = (
        df.groupby(["Dept", "Class", "Subclass"], as_index=False)
        .agg(
            dtc_netsales=("DTC Netsales $s", "sum"),
            dtc_netsales_ly=("DTC Netsales $s LY", "sum"),
            dtc_demand=(demand_col, "sum"),
            sku_count=("SKU", "nunique"),
        )
    )

    summary["sales_yoy_pct"] = (
        (summary["dtc_netsales"] - summary["dtc_netsales_ly"])
        .div(summary["dtc_netsales_ly"].replace(0, pd.NA))
        * 100
    ).astype("Float64").round(1)


    return summary.sort_values("dtc_netsales", ascending=False)


def build_wow_summary(
    latest_summary: pd.DataFrame,
    prior_df: pd.DataFrame,
) -> pd.DataFrame:
    """Build week-over-week category comparison."""
    prior_summary = (
        prior_df.groupby(["Dept", "Class", "Subclass"], as_index=False)
        .agg(dtc_netsales_pw=("DTC Netsales $s", "sum"))
    )

    wow = pd.merge(
        latest_summary,
        prior_summary,
        on=["Dept", "Class", "Subclass"],
        how="left",
    )

    wow["wow_pct"] = (
        (wow["dtc_netsales"] - wow["dtc_netsales_pw"])
        .div(wow["dtc_netsales_pw"].replace(0, pd.NA))
        * 100
    ).astype("Float64").round(1)

    return wow.sort_values("dtc_netsales", ascending=False)


def build_top_items(df: pd.DataFrame, n: int = 10) -> pd.DataFrame:
    """Return top N items by DTC Netsales."""
    top_items = (
        df.groupby(["SKU", "Item"], as_index=False)["DTC Netsales $s"]
        .sum()
        .sort_values("DTC Netsales $s", ascending=False)
        .head(n)
    )
    return top_items


def build_overview(
    latest_week: pd.Timestamp,
    prior_week: pd.Timestamp,
    df_latest: pd.DataFrame,
    df_prior: pd.DataFrame,
    latest_summary: pd.DataFrame,
    top_items_latest: pd.DataFrame,
) -> pd.DataFrame:
    """Build a simple executive overview table."""
    latest_sales = df_latest["DTC Netsales $s"].sum()
    prior_sales = df_prior["DTC Netsales $s"].sum()

    wow_pct = (
        ((latest_sales - prior_sales) / prior_sales) * 100
        if prior_sales != 0
        else pd.NA
    )

    top_subclass = latest_summary.iloc[0]["Subclass"] if not latest_summary.empty else pd.NA
    top_sku = top_items_latest.iloc[0]["SKU"] if not top_items_latest.empty else pd.NA

    overview = pd.DataFrame(
        {
            "Metric": [
                "Latest Week End",
                "Prior Week End",
                "Latest Week Sales",
                "Prior Week Sales",
                "WoW %",
                "Top Subclass (Latest Week)",
                "Top SKU (Latest Week)",
            ],
            "Value": [
                latest_week,
                prior_week,
                latest_sales,
                prior_sales,
                wow_pct,
                top_subclass,
                top_sku,
            ],
        }
    )

    return overview


def export_report(output_path: str, outputs: dict[str, pd.DataFrame]) -> None:
    """Export all outputs into one formatted Excel workbook."""
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        for sheet_name, data in outputs.items():
            safe_sheet_name = sheet_name[:31]
            data.to_excel(writer, sheet_name=safe_sheet_name, index=False)

            ws = writer.sheets[safe_sheet_name]

            # Freeze header row
            ws.freeze_panes = "A2"

            # Header formatting
            header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
            header_font = Font(color="FFFFFF", bold=True)
            header_alignment = Alignment(horizontal="center", vertical="center")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            # Detect columns by name for formatting
            currency_keywords = ["sales", "netsales", "demand"]
            percent_keywords = ["pct", "%"]
            date_keywords = ["week", "date"]

            # Auto-width + format columns
            for col_idx, column_name in enumerate(data.columns, start=1):
                col_letter = get_column_letter(col_idx)

                # Set width based on longest value in the column
                max_length = len(str(column_name))
                for value in data[column_name]:
                    if pd.notna(value):
                        max_length = max(max_length, len(str(value)))
                ws.column_dimensions[col_letter].width = min(max_length + 2, 35)

                col_name_lower = column_name.lower()

                for row_idx in range(2, len(data) + 2):
                    cell = ws[f"{col_letter}{row_idx}"]

                    # Currency formatting
                    if any(keyword in col_name_lower for keyword in currency_keywords):
                        if pd.notna(cell.value) and isinstance(cell.value, (int, float)):
                            cell.number_format = '$#,##0'

                    # Percent formatting
                    elif any(keyword in col_name_lower for keyword in percent_keywords):
                        if pd.notna(cell.value) and isinstance(cell.value, (int, float)):
                            cell.number_format = '0.0'

                    # Date formatting
                    elif any(keyword in col_name_lower for keyword in date_keywords):
                        if hasattr(cell.value, "year"):
                            cell.number_format = 'yyyy-mm-dd'

            # Optional: turn on autofilter
            ws.auto_filter.ref = ws.dimensions
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            # Find YoY column index
            for col_idx, col_name in enumerate(data.columns, start=1):
                col_name_lower = col_name.lower()
   
                if "yoy" in col_name_lower or "wow" in col_name_lower or "pct" in col_name_lower:
                    col_letter = get_column_letter(col_idx)
                    start_row = 2
                    end_row = len(data) + 1
                    range_str = f"{col_letter}{start_row}:{col_letter}{end_row}"

                    ws.conditional_formatting.add(
                        range_str,
                        CellIsRule(
                            operator='greaterThan', 
                            formula=['0'], 
                            fill=green_fill,
                            stopIfTrue=True
                        ),
                    )
                    ws.conditional_formatting.add(
                        range_str,
                        CellIsRule(operator='lessThan', 
                               formula=['0'], 
                               fill=red_fill,
                               stopIfTrue=True,
                        ),
                    )
            if safe_sheet_name == "Overview":
                ws.column_dimensions["A"].width = 28
                ws.column_dimensions["B"].width = 18
                for row in range(2, len(data) + 2):
                    ws[f"A{row}"].font = Font(bold=True)     

def parse_args():
    parser = argparse.ArgumentParser(
        description="Build an inventory analysis Excel report."
    )

    parser.add_argument(
        "--input",
        default="../data/dummy revenue data.xlsx",
        help="Path to the input Excel file"
    )

    parser.add_argument(
        "--output",
        default="../outputs/inventory_report.xlsx",
        help="Path to the output Excel report"
    )

    parser.add_argument(
        "--top-n",
        type=int,
        default=10,
        help="Number of top items to return"
    )

    parser.add_argument(
        "--dept",
        type=str,
        default=None,
        help="Optional department filter, e.g. 914"
    )

    return parser.parse_args()

def main() -> None:
    args = parse_args()

    input_file = args.input
    output_file = args.output
    top_n = args.top_n
    dept_filter = args.dept

    # Load + clean
    df = load_data(input_file)
    df = clean_data(df)
    validate_columns(df)
    demand_col = get_demand_column(df)

    # Optional department filter
    if dept_filter is not None:
        df = df[df["Dept"].astype(str) == str(dept_filter)].copy()
        print(f"\nApplied Dept filter: {dept_filter}")
        print("Filtered shape:", df.shape)

        if df.empty:
            raise ValueError(f"No rows found for Dept {dept_filter}")

    # Latest + prior week
    df_latest, latest_week = get_latest_week(df)
    df_prior, prior_week = get_prior_week(df, latest_week)

    print("\nLatest Week End:", latest_week)
    print("Latest Week Shape:", df_latest.shape)
    print(df_latest[["Week Start", "Week End"]].drop_duplicates().head())

    print("\nPrior Week:", prior_week)
    print("Prior Week Shape:", df_prior.shape)

    # Build outputs
    category_summary_latest = build_category_summary(df_latest, demand_col)
    category_summary_full = build_category_summary(df, demand_col)
    category_wow = build_wow_summary(category_summary_latest, df_prior)
    top_items_latest = build_top_items(df_latest, n=top_n)
    top_items_full = build_top_items(df, n=top_n)
    overview = build_overview(
        latest_week,
        prior_week,
        df_latest,
        df_prior,
        category_summary_latest,
        top_items_latest,
    )

    # Print previews
    print("\nCategory Summary - Latest Week")
    print(category_summary_latest.head(10))

    print("\nCategory Summary - Full Period")
    print(category_summary_full.head(10))

    print("\nCategory WoW Summary")
    print(category_wow.head(10))

    print(f"\nTop {top_n} Items by DTC Netsales - Latest Week")
    print(top_items_latest)

    print(f"\nTop {top_n} Items by DTC Netsales - Full Period")
    print(top_items_full)

    outputs = {
        "Overview": overview,
        "Latest Week Summary": category_summary_latest,
        "Full Category Summary": category_summary_full,
        "WoW Summary": category_wow,
        "Top Items Latest": top_items_latest,
        "Top Items Full": top_items_full,
    }

    export_report(output_file, outputs)
    print(f"\nReport created: {output_file}")
    
if __name__ == "__main__":
    main()